import customtkinter as ctk
from tkinter import messagebox
import datetime
import openpyxl

ctk.set_appearance_mode("dark")  # Modes: system (default), light, dark
ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

class UserDataEntry:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("User Data Entry")
        self.root.geometry("400x500")

        self.create_widgets()

    def create_widgets(self):
        # Main frame
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(pady=20, padx=20, fill="both", expand=True)

        # Title
        title = ctk.CTkLabel(main_frame, text="User Data Entry", font=("Roboto", 24))
        title.pack(pady=10)

        # Name
        self.name_entry = ctk.CTkEntry(main_frame, placeholder_text="Name")
        self.name_entry.pack(pady=10, padx=20, fill="x")

        # Age
        self.age_entry = ctk.CTkEntry(main_frame, placeholder_text="Age")
        self.age_entry.pack(pady=10, padx=20, fill="x")

        # Gender
        self.gender_var = ctk.StringVar(value="Male")
        gender_frame = ctk.CTkFrame(main_frame)
        gender_frame.pack(pady=10, padx=20, fill="x")
        gender_label = ctk.CTkLabel(gender_frame, text="Gender:")
        gender_label.pack(side="left", padx=(0, 10))
        male_radio = ctk.CTkRadioButton(gender_frame, text="Male", variable=self.gender_var, value="Male")
        male_radio.pack(side="left", padx=(0, 10))
        female_radio = ctk.CTkRadioButton(gender_frame, text="Female", variable=self.gender_var, value="Female")
        female_radio.pack(side="left")

        # Qualification
        self.qualification_entry = ctk.CTkEntry(main_frame, placeholder_text="Qualification")
        self.qualification_entry.pack(pady=10, padx=20, fill="x")

        # Submit button
        submit_button = ctk.CTkButton(main_frame, text="Submit", command=self.submit_data)
        submit_button.pack(pady=10, padx=20, fill="x")

        # Extract button
        extract_button = ctk.CTkButton(main_frame, text="Extract to Excel", command=self.extract_to_excel)
        extract_button.pack(pady=10, padx=20, fill="x")

    def submit_data(self):
        name = self.name_entry.get()
        age = self.age_entry.get()
        gender = self.gender_var.get()
        qualification = self.qualification_entry.get()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Append data to a text file
        with open("user_data.txt", "a") as file:
            file.write(f"{name},{age},{gender},{qualification},{timestamp}\n")

        messagebox.showinfo("Success", "Data submitted successfully!")

        # Clear the entries
        self.name_entry.delete(0, ctk.END)
        self.age_entry.delete(0, ctk.END)
        self.qualification_entry.delete(0, ctk.END)

    def extract_to_excel(self):
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "User Data"

        # Add headers
        headers = ["Name", "Age", "Gender", "Qualification", "Timestamp"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Read data from the text file and add to the sheet
        with open("user_data.txt", "r") as file:
            for row, line in enumerate(file, start=2):
                data = line.strip().split(",")
                for col, value in enumerate(data, start=1):
                    sheet.cell(row=row, column=col, value=value)

        # Save the workbook
        wb.save("user_data.xlsx")
        messagebox.showinfo("Success", "Data extracted to user_data.xlsx successfully!")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = UserDataEntry()
    app.run()